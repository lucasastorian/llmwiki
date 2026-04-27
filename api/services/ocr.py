import json
import base64
import asyncio
import logging
import subprocess
import tempfile
from pathlib import Path

import asyncpg
import httpx

from config import settings
from services.s3 import S3Service
from services.chunker import chunk_text, chunk_pages, store_chunks
from services.pdf_extract import extract_pdf

logger = logging.getLogger(__name__)

MISTRAL_OCR_URL = "https://api.mistral.ai/v1/ocr"
MAX_RETRIES = 3
RETRY_BACKOFF = [2, 5, 10]

OFFICE_TYPES = {"pptx", "ppt", "docx", "doc"}
IMAGE_TYPES = {"png", "jpg", "jpeg", "webp", "gif"}
OCR_TYPES = {"pdf"} | OFFICE_TYPES | IMAGE_TYPES

class OCRService:
    def __init__(self, s3: S3Service, pool: asyncpg.Pool):
        self._s3 = s3
        self._pool = pool
        self._semaphore = asyncio.Semaphore(3)

    async def process_document(self, document_id: str, user_id: str):
        async with self._semaphore:
            await self._do_process(document_id, user_id)

    async def _check_global_limits(self, document_id: str):
        if not settings.GLOBAL_OCR_ENABLED:
            raise ValueError("OCR processing is temporarily disabled by the administrator.")

        total_pages = await self._pool.fetchval(
            "SELECT COALESCE(SUM(page_count), 0) FROM documents WHERE NOT archived"
        )
        if total_pages >= settings.GLOBAL_MAX_PAGES:
            raise ValueError(
                f"Platform page limit reached ({settings.GLOBAL_MAX_PAGES:,} pages). "
                "Please contact the administrator."
            )

    async def _do_process(self, document_id: str, user_id: str):
        try:
            await self._check_global_limits(document_id)
            await self._set_status(document_id, "processing")

            doc = await self._pool.fetchrow(
                "SELECT filename, file_type, knowledge_base_id::text as kb_id "
                "FROM documents WHERE id = $1 AND user_id = $2",
                document_id, user_id,
            )
            if not doc:
                logger.error("Document %s not found for user %s", document_id, user_id)
                return

            ext = doc["filename"].rsplit(".", 1)[-1].lower() if "." in doc["filename"] else doc["file_type"]
            kb_id = doc["kb_id"]
            s3_source_key = f"{user_id}/{document_id}/source.{ext}"

            if ext in OFFICE_TYPES:
                await self._process_office(document_id, user_id, kb_id, s3_source_key, ext)
            elif ext in IMAGE_TYPES:
                await self._process_image(document_id, user_id, s3_source_key, ext)
            elif ext == "pdf":
                await self._process_pdf(document_id, user_id, kb_id, s3_source_key)
            elif ext in ("html", "htm"):
                await self._process_html(document_id, user_id, kb_id, s3_source_key)
            elif ext in ("xlsx", "xls", "csv"):
                await self._process_spreadsheet(document_id, user_id, kb_id, s3_source_key, ext)
            else:
                raise ValueError(f"Unsupported file type: {ext}")

        except Exception as e:
            logger.exception("Processing failed for document %s", document_id)
            try:
                await self._pool.execute(
                    "UPDATE documents SET status = 'failed', error_message = $2, updated_at = now() "
                    "WHERE id = $1",
                    document_id, str(e)[:500],
                )
            except Exception:
                logger.exception("Failed to update status to failed for %s", document_id)

    # ── PDF extraction ────────────────────────────────────────────────────

    async def _process_pdf(self, document_id: str, user_id: str, kb_id: str, s3_source_key: str):
        if settings.PDF_BACKEND == "mistral":
            if not settings.MISTRAL_API_KEY:
                raise ValueError("MISTRAL_API_KEY not configured — cannot process PDFs")
            presigned_url = await self._s3.generate_presigned_get(s3_source_key)
            ocr_result = await self._call_mistral_ocr(presigned_url, "document_url")
            await self._store_ocr_result(document_id, user_id, kb_id, ocr_result)
        elif settings.CONVERTER_URL:
            presigned_url = await self._s3.generate_presigned_get(s3_source_key)
            pages = await self._call_converter_extract(presigned_url, "pdf")
            await self._store_extracted_pages(document_id, user_id, kb_id, pages, "opendataloader")
        else:
            await self._process_opendataloader(document_id, user_id, kb_id, s3_source_key)

    async def _process_office(self, document_id: str, user_id: str, kb_id: str, s3_source_key: str, ext: str):
        """Process Office files. Routes through converter or falls back to local LibreOffice."""
        if settings.PDF_BACKEND == "mistral":
            pdf_key = await self._convert_to_pdf_s3(document_id, user_id, s3_source_key, ext)
            if not settings.MISTRAL_API_KEY:
                raise ValueError("MISTRAL_API_KEY not configured")
            presigned_url = await self._s3.generate_presigned_get(pdf_key)
            ocr_result = await self._call_mistral_ocr(presigned_url, "document_url")
            await self._store_ocr_result(document_id, user_id, kb_id, ocr_result)
        elif settings.CONVERTER_URL:
            presigned_url = await self._s3.generate_presigned_get(s3_source_key)
            pages = await self._call_converter_extract(presigned_url, ext)
            await self._store_extracted_pages(document_id, user_id, kb_id, pages, "opendataloader")
        else:
            await self._process_office_local(document_id, user_id, kb_id, s3_source_key, ext)

    # ── Converter integration (hosted mode) ───────────────────────────────

    async def _call_converter_extract(self, source_url: str, ext: str) -> list[tuple[int, str]]:
        """Call the converter /extract endpoint. Returns list of (page_num, markdown).

        Sends a request_id for source binding. If the converter echoes it back,
        we verify the match; if the converter doesn't support it, we log a warning
        but still accept the response (forward-compatible).
        """
        import uuid as _uuid
        request_id = str(_uuid.uuid4())

        headers = {}
        if settings.CONVERTER_SECRET:
            headers["Authorization"] = f"Bearer {settings.CONVERTER_SECRET}"

        async with httpx.AsyncClient(timeout=httpx.Timeout(300.0, connect=10.0)) as client:
            resp = await client.post(
                f"{settings.CONVERTER_URL}/extract",
                json={"source_url": source_url, "source_ext": ext, "request_id": request_id},
                headers=headers,
            )
            resp.raise_for_status()
            data = resp.json()

        # Source binding: verify request_id echo if the converter supports it
        echoed_id = data.get("request_id")
        if echoed_id is not None and echoed_id != request_id:
            raise ValueError(
                f"Converter response binding mismatch: sent {request_id}, got {echoed_id}. "
                "Possible stale cache or cross-request contamination."
            )
        if echoed_id is None:
            logger.warning("Converter did not echo request_id — source binding not verified")

        pages = data.get("pages", [])
        if not pages:
            raise ValueError("Converter returned empty pages — extraction may have failed silently")

        return [(p["page"], p["content"]) for p in pages]

    # ── OpenDataLoader local extraction ───────────────────────────────────

    async def _process_opendataloader(self, document_id: str, user_id: str, kb_id: str, s3_source_key: str):
        """Extract PDF via opendataloader-pdf (local mode or hosted fallback)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            pdf_path = Path(tmpdir) / "source.pdf"
            await self._s3.download_to_file(s3_source_key, str(pdf_path))
            pages_with_images = await asyncio.to_thread(extract_pdf, str(pdf_path))

        # Upload extracted images to S3 and build per-page elements metadata
        page_elements: dict[int, dict] = {}
        for page_num, _, images in pages_with_images:
            if not images:
                continue
            page_imgs = []
            for img in images:
                mime = "image/jpeg" if img["format"] == "jpeg" else "image/png"
                await self._s3.upload_bytes(
                    f"{user_id}/{document_id}/images/{img['id']}",
                    img["bytes"], mime,
                )
                page_imgs.append({"id": img["id"]})
            page_elements[page_num] = {"images": page_imgs}

        page_contents = [(num, md) for num, md, _ in pages_with_images]
        await self._store_extracted_pages(
            document_id, user_id, kb_id, page_contents, "opendataloader",
            page_elements=page_elements,
        )

    # ── Office local fallback (no converter) ──────────────────────────────

    async def _convert_to_pdf_s3(self, document_id: str, user_id: str, s3_source_key: str, ext: str) -> str:
        """Convert Office file to PDF and upload to S3. Returns S3 key of the PDF."""
        pdf_key = f"{user_id}/{document_id}/converted.pdf"

        if settings.CONVERTER_URL:
            # Legacy path — only used for Mistral backend with converter
            import uuid as _uuid
            request_id = str(_uuid.uuid4())
            source_url = await self._s3.generate_presigned_get(s3_source_key)
            result_url = await self._s3.generate_presigned_put(pdf_key)
            headers = {}
            if settings.CONVERTER_SECRET:
                headers["Authorization"] = f"Bearer {settings.CONVERTER_SECRET}"
            async with httpx.AsyncClient(timeout=httpx.Timeout(180.0, connect=10.0)) as client:
                resp = await client.post(
                    f"{settings.CONVERTER_URL}/convert",
                    json={
                        "source_url": source_url, "result_url": result_url,
                        "source_ext": ext, "request_id": request_id,
                    },
                    headers=headers,
                )
                resp.raise_for_status()
                data = resp.json() if resp.headers.get("content-type", "").startswith("application/json") else {}
                echoed_id = data.get("request_id") if isinstance(data, dict) else None
                if echoed_id is not None and echoed_id != request_id:
                    raise ValueError(
                        f"Converter response binding mismatch: sent {request_id}, got {echoed_id}"
                    )
        else:
            with tempfile.TemporaryDirectory() as tmpdir:
                source_path = Path(tmpdir) / f"source.{ext}"
                await self._s3.download_to_file(s3_source_key, str(source_path))
                result = await asyncio.to_thread(
                    subprocess.run,
                    ["libreoffice", "--headless", "--norestore", "--convert-to", "pdf", "--outdir", tmpdir, str(source_path)],
                    capture_output=True, timeout=120,
                )
                if result.returncode != 0:
                    raise RuntimeError(f"LibreOffice conversion failed: {result.stderr.decode()[:300]}")
                pdf_path = Path(tmpdir) / "source.pdf"
                if not pdf_path.exists():
                    raise RuntimeError("LibreOffice did not produce a PDF")
                await self._s3.upload_file(pdf_key, str(pdf_path), "application/pdf")

        return pdf_key

    async def _process_office_local(self, document_id: str, user_id: str, kb_id: str, s3_source_key: str, ext: str):
        """Convert Office file to PDF locally, then extract with opendataloader."""
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = Path(tmpdir) / f"source.{ext}"
            await self._s3.download_to_file(s3_source_key, str(source_path))

            result = await asyncio.to_thread(
                subprocess.run,
                ["libreoffice", "--headless", "--norestore", "--convert-to", "pdf", "--outdir", tmpdir, str(source_path)],
                capture_output=True, timeout=120,
            )
            if result.returncode != 0:
                raise RuntimeError(f"LibreOffice conversion failed: {result.stderr.decode()[:300]}")

            pdf_path = Path(tmpdir) / "source.pdf"
            if not pdf_path.exists():
                raise RuntimeError("LibreOffice did not produce a PDF")

            pages_with_images = await asyncio.to_thread(extract_pdf, str(pdf_path))

        page_elements: dict[int, dict] = {}
        for page_num, _, images in pages_with_images:
            if not images:
                continue
            page_imgs = []
            for img in images:
                mime = "image/jpeg" if img["format"] == "jpeg" else "image/png"
                await self._s3.upload_bytes(
                    f"{user_id}/{document_id}/images/{img['id']}",
                    img["bytes"], mime,
                )
                page_imgs.append({"id": img["id"]})
            page_elements[page_num] = {"images": page_imgs}

        page_contents = [(num, md) for num, md, _ in pages_with_images]
        await self._store_extracted_pages(
            document_id, user_id, kb_id, page_contents, "libreoffice+opendataloader",
            page_elements=page_elements,
        )

    # ── Shared page storage ───────────────────────────────────────────────

    async def _store_extracted_pages(
        self, document_id: str, user_id: str, kb_id: str,
        page_contents: list[tuple[int, str]], parser: str,
        page_elements: dict[int, dict] | None = None,
    ):
        """Store pages/chunks and update document status."""
        num_pages = len(page_contents)

        if num_pages > settings.QUOTA_MAX_PAGES_PER_DOC:
            raise ValueError(
                f"Document has {num_pages} pages, maximum is {settings.QUOTA_MAX_PAGES_PER_DOC}."
            )

        conn = await self._pool.acquire()
        try:
            await conn.execute("DELETE FROM document_pages WHERE document_id = $1", document_id)
            for num, md in page_contents:
                elements = (page_elements or {}).get(num)
                await conn.execute(
                    "INSERT INTO document_pages (document_id, page, content, elements) "
                    "VALUES ($1, $2, $3, $4)",
                    document_id, num, md,
                    json.dumps(elements) if elements else None,
                )
        finally:
            await self._pool.release(conn)

        full_content = "\n\n---\n\n".join(md for _, md in page_contents)
        chunks = chunk_pages(page_contents)
        await store_chunks(self._pool, document_id, user_id, kb_id, chunks)

        await self._pool.execute(
            "UPDATE documents SET status = 'ready', content = $2, page_count = $3, "
            "parser = $4, updated_at = now() WHERE id = $1",
            document_id, full_content, num_pages, parser,
        )
        logger.info("Extracted (%s): doc=%s pages=%d chunks=%d", parser, document_id[:8], num_pages, len(chunks))

    # ── Image processing ──────────────────────────────────────────────────

    async def _process_image(self, document_id: str, user_id: str, s3_source_key: str, ext: str):
        """Images are stored as-is. No OCR. The MCP read tool returns them natively."""
        await self._pool.execute(
            "UPDATE documents SET status = 'ready', page_count = 1, parser = 'native', updated_at = now() "
            "WHERE id = $1",
            document_id,
        )
        logger.info("Image stored: doc=%s", document_id[:8])

    # ── HTML processing ───────────────────────────────────────────────────

    async def _process_html(self, document_id: str, user_id: str, kb_id: str, s3_source_key: str):
        """Parse HTML with webmd parser, store markdown + tagged HTML."""
        from html_parser import Parser

        html_bytes = await self._s3.download_bytes(s3_source_key)
        raw_html = html_bytes.decode("utf-8", errors="replace")

        parser = Parser(raw_html, content_only=True)
        result = parser.parse()

        await parser.embed_images()
        tagged_html = parser.html()

        await self._s3.upload_bytes(
            f"{user_id}/{document_id}/tagged.html",
            tagged_html.encode("utf-8"),
            "text/html",
        )

        markdown_content = result.content
        chunks = chunk_text(markdown_content)
        await store_chunks(self._pool, document_id, user_id, kb_id, chunks)

        await self._pool.execute(
            "UPDATE documents SET status = 'ready', content = $2, page_count = 1, parser = 'webmd', updated_at = now() "
            "WHERE id = $1",
            document_id, markdown_content,
        )
        logger.info("HTML processed: doc=%s chunks=%d", document_id[:8], len(chunks))

    # ── Spreadsheet processing ────────────────────────────────────────────

    async def _process_spreadsheet(self, document_id: str, user_id: str, kb_id: str, s3_source_key: str, ext: str):
        """Download spreadsheet, store each sheet as a document_page."""
        with tempfile.TemporaryDirectory() as tmpdir:
            source_path = Path(tmpdir) / f"source.{ext}"
            await self._s3.download_to_file(s3_source_key, str(source_path))

            sheets = await asyncio.to_thread(self._parse_sheets, str(source_path), ext)

            conn = await self._pool.acquire()
            try:
                await conn.execute("DELETE FROM document_pages WHERE document_id = $1", document_id)
                content_parts = []
                for i, (name, md) in enumerate(sheets, 1):
                    content_parts.append(f"## {name}\n\n{md}")
                    await conn.execute(
                        "INSERT INTO document_pages (document_id, page, content, elements) "
                        "VALUES ($1, $2, $3, $4)",
                        document_id, i, md,
                        json.dumps({"sheet_name": name}),
                    )
            finally:
                await self._pool.release(conn)

            full_content = "\n\n---\n\n".join(content_parts)
            page_contents = [(i + 1, md) for i, (_, md) in enumerate(sheets)]
            chunks = chunk_pages(page_contents)
            await store_chunks(self._pool, document_id, user_id, kb_id, chunks)

            await self._pool.execute(
                "UPDATE documents SET status = 'ready', content = $2, page_count = $3, parser = 'openpyxl', updated_at = now() "
                "WHERE id = $1",
                document_id, full_content, len(sheets),
            )
            logger.info("Spreadsheet processed: doc=%s sheets=%d chunks=%d", document_id[:8], len(sheets), len(chunks))

    @staticmethod
    def _rows_to_markdown(rows: list[list[str]], max_rows: int = 100) -> str:
        if not rows:
            return "(empty)"
        header = "| " + " | ".join(rows[0]) + " |"
        sep = "| " + " | ".join("---" for _ in rows[0]) + " |"
        data = rows[1:max_rows + 1]
        body = "\n".join("| " + " | ".join(r) + " |" for r in data)
        truncated = f"\n\n*({len(rows) - 1 - max_rows} more rows truncated)*" if len(rows) - 1 > max_rows else ""
        return f"{header}\n{sep}\n{body}{truncated}"

    @staticmethod
    def _parse_sheets(path: str, ext: str) -> list[tuple[str, str]]:
        """Returns list of (sheet_name, markdown_table) tuples."""
        import csv
        if ext == "csv":
            with open(path, newline="", encoding="utf-8", errors="replace") as f:
                rows = [[c for c in row] for row in csv.reader(f)]
            return [("Sheet1", OCRService._rows_to_markdown(rows))]

        try:
            import openpyxl
        except ImportError:
            return [("Error", "(openpyxl not installed)")]

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = [[str(cell) if cell is not None else "" for cell in row] for row in ws.iter_rows(values_only=True)]
            if not rows:
                continue
            sheets.append((name, OCRService._rows_to_markdown(rows)))
        wb.close()
        return sheets

    # ── Mistral OCR ───────────────────────────────────────────────────────

    async def _store_ocr_result(self, document_id: str, user_id: str, kb_id: str, ocr_result: dict):
        ocr_json_bytes = json.dumps(ocr_result).encode()
        await self._s3.upload_bytes(f"{user_id}/{document_id}/ocr.json", ocr_json_bytes, "application/json")

        pages = ocr_result.get("pages", [])

        if len(pages) > settings.QUOTA_MAX_PAGES_PER_DOC:
            raise ValueError(
                f"Document has {len(pages)} pages, maximum is {settings.QUOTA_MAX_PAGES_PER_DOC}."
            )

        for page in pages:
            for img in page.get("images", []):
                img_id = img.get("id")
                img_b64 = img.get("image_base64")
                if not img_id or not img_b64:
                    continue
                if img_b64.startswith("data:"):
                    img_b64 = img_b64.split(",", 1)[1]
                img_bytes = base64.b64decode(img_b64)
                await self._s3.upload_bytes(
                    f"{user_id}/{document_id}/images/{img_id}",
                    img_bytes,
                    "image/jpeg",
                )

        content_parts = []
        conn = await self._pool.acquire()
        try:
            await conn.execute("DELETE FROM document_pages WHERE document_id = $1", document_id)
            for page in pages:
                page_index = page.get("index", 0) + 1
                page_md = page.get("markdown", "")
                content_parts.append(page_md)

                elements = {}
                if page.get("images"):
                    elements["images"] = [
                        {k: v for k, v in img.items() if k != "image_base64"}
                        for img in page["images"]
                    ]
                if page.get("dimensions"):
                    elements["dimensions"] = page["dimensions"]
                if page.get("tables"):
                    elements["tables"] = page["tables"]

                await conn.execute(
                    "INSERT INTO document_pages (document_id, page, content, elements) "
                    "VALUES ($1, $2, $3, $4)",
                    document_id, page_index, page_md,
                    json.dumps(elements) if elements else None,
                )
        finally:
            await self._pool.release(conn)

        full_content = "\n\n---\n\n".join(content_parts)
        page_count = len(pages)

        page_contents = [(page.get("index", 0) + 1, page.get("markdown", "")) for page in pages]
        chunks = chunk_pages(page_contents)
        await store_chunks(self._pool, document_id, user_id, kb_id, chunks)

        await self._pool.execute(
            "UPDATE documents SET status = 'ready', content = $2, page_count = $3, parser = 'mistral', updated_at = now() "
            "WHERE id = $1",
            document_id, full_content, page_count,
        )
        logger.info("OCR complete: doc=%s pages=%d chunks=%d", document_id[:8], page_count, len(chunks))

    async def _call_mistral_ocr(self, url: str, url_type: str = "document_url") -> dict:
        last_error = None
        for attempt in range(MAX_RETRIES):
            try:
                async with httpx.AsyncClient(timeout=httpx.Timeout(300.0, connect=10.0)) as client:
                    resp = await client.post(
                        MISTRAL_OCR_URL,
                        headers={
                            "Authorization": f"Bearer {settings.MISTRAL_API_KEY}",
                            "Content-Type": "application/json",
                        },
                        json={
                            "model": "mistral-ocr-latest",
                            "document": {
                                "type": url_type,
                                url_type: url,
                            },
                            "include_image_base64": True,
                            "table_format": "markdown",
                        },
                    )
                    resp.raise_for_status()
                    return resp.json()
            except (httpx.HTTPStatusError, httpx.TimeoutException) as e:
                last_error = e
                if attempt < MAX_RETRIES - 1:
                    wait = RETRY_BACKOFF[attempt]
                    logger.warning("Mistral OCR attempt %d failed: %s, retrying in %ds", attempt + 1, e, wait)
                    await asyncio.sleep(wait)
        raise last_error or RuntimeError("Mistral OCR failed after retries")

    async def _set_status(self, document_id: str, status: str):
        await self._pool.execute(
            "UPDATE documents SET status = $2, updated_at = now() WHERE id = $1",
            document_id, status,
        )
